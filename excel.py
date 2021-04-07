import os
import json
from openpyxl import Workbook
from openpyxl import load_workbook

class AdInfo:
    
    def __init__(self, adinfo_template):
        '''
        :param: adinfo_template: filename for adinfo json template.

        '''
        self.adinfo_template = adinfo_template or 'adinfo.json'

    def _open_file(self, filename):
        print(filename)
        wb = load_workbook(filename=filename)
        return wb

    def adinfo_serializer(self, adinfo_filename):
        """
        Get adinfo filename and template and creates a dictionary with
        all the adinfo data.
        This will not work for formulas cells.

        :adinfo_filename: The excel adinfo filename.

        return a dictionary version of the adinfo file.
        """
        """
        If formula cells are needed, should use:
        import xlwings as xw
        wbxl=xw.Book('to_erase.xlsx')
        wbxl.sheets['Sheet'].range('B3').value

        """
        adinfo_data = self._open_file(adinfo_filename).active
        template_dict = dict()
        adinfo_dict = dict(heading={}, lines={})
        
        with open(self.adinfo_template, 'r') as template:
            template_dict = json.load(template)
        
        for key, value in template_dict['heading'].items():
            adinfo_dict['heading'][key] = adinfo_data[value].value

        initial_row = (template_dict['lines']).pop('linha_inicial') + 1
        final_row = adinfo_data.max_row

        for row in range(initial_row, final_row):
            adinfo_dict['lines'][row] = dict()
            for value_name, column in template_dict['lines'].items():
                adinfo_dict['lines'][row][value_name] = adinfo_data[
                    column+str(row)].value
        return adinfo_dict
"""
path = os.path.dirname(os.path.realpath(__file__))
adinfo = AdInfo(path+'/adinfo/adinfo.json')
adinfo_dict = adinfo.adinfo_serializer(path+'/adinfo/adinfo2.xlsx')
print (adinfo_dict)
"""
